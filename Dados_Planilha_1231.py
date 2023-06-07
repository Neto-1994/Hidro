from sys import displayhook
import pandas
import conexao
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Busca de dados no banco
try:
    class Estacao1():
        def _init_(self, data1, data2, Nome_Arquivo, Nome_Salvar):
            data2 = data2 + " 21:00:00"
            consulta_sql = "select date(Horalocal), sum(Pluvio1h), sum(Evaporacao1h) from medicoes where codigo_sec = 1231 and horalocal between %s and %s group by date(Horalocal);"
            cursor = conexao.con.cursor()
            cursor.execute(consulta_sql, (data1, data2))
            Dados = cursor.fetchall()

# Gerar dataframe com os dados
            df = pandas.DataFrame(
                Dados, columns=["DATA", "PRECIP. (mm)", "EVAPORACAO (mm)"])

# Condicao acumulativa da precipitacao mensal e insercao do valor no dataframe
            Acumulativo = df["PRECIP. (mm)"].cumsum()
            df.insert(2, "PREC. ACUM. (mm)", Acumulativo)

# Formatacao da data
            df["DATA"] = pandas.to_datetime(df.DATA)
            df["DATA"] = df["DATA"].dt.strftime("%m/%d/%y")

# Carregar arquivo excel existente
            wb = load_workbook("C:/"+Nome_Arquivo+".xlsx")
            ws = wb["Evapo 1"]

# Inserir dados na planilha
            dr = dataframe_to_rows(df, index=False, header=False,)
            for r in dr:
                ws.append(r)

# Inserir imagens na planilha
#            img1 = Image("C:/Users/Jair/Pictures/Acqua.png")
#            img2 = Image("C:/Users/Jair/Pictures/Lundin.png")

#            ws.add_image(img1, "A1")
#            ws.add_image(img2, "D1")

# Formatar dados da planilha
            row = ws.max_row
            column = ws.max_column
            for i in range(343, row + 1):
                for j in range(1, column):
                    ws.cell(i, j).font = Font(name="Times New Roman",
                                              size=12)
#                                             bold = False,
#                                             italic = False,
                    ws.cell(i, j).border = Border(left=Side(border_style="thin",
                                                            color='FF000000'),
                                                  right=Side(border_style="thin",
                                                             color='FF000000'),
                                                  top=Side(border_style="thin",
                                                           color='FF000000'),
                                                  bottom=Side(border_style="thin",
                                                              color='FF000000'))
#                                                 diagonal=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 diagonal_direction=0,
#                                                 outline=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 vertical=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 horizontal=Side(border_style=None,
#                                                 color='FF000000'))

                    ws.cell(i, j).alignment = Alignment(horizontal='center', vertical='center')
                    ws.cell(i, j).number_format = '0.0'

# Apresentacao dos dataframes no terminal
#               displayhook(df)

# Exportar dataframes como arquivo xlsx
#               df.to_excel("Teste Salvamento.xlsx", index= False) # Gerar arquivo pelo pandas
            wb.save(Nome_Salvar+".xlsx")  # Gerar arquivo pelo openpyxl
            print("\nArquivo excel criado com sucesso!!!\n")
            
except OSError as e:
    print("Erro: ", e)
