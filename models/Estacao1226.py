from sys import displayhook
import pandas
import conexao
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image

# Busca de dados no banco
try:
    class Estacao1226():
        def _init_(self, data1, data2, Nome_Arquivo, Nome_Salvar):
            consulta_sql = "SELECT DATE(HoraLocal), AVG(SPressao), AVG(Vazao_calculada) FROM medicoes WHERE Codigo_Sec = 1226 AND HoraLocal between %s and %s GROUP BY DATE(HoraLocal);"
            cursor = conexao.con.cursor()
            cursor.execute(consulta_sql, (data1, data2))
            Dados = cursor.fetchall()

# Gerar dataframe com os dados
            df = pandas.DataFrame(Dados, columns=["Data", "Nível Médio Diário (m)", "Vazão Média Diária(m³/s)"])

# Cálculo da Vazão Média por segundo e inserção do valor no dataframe
            horario = df["Vazão Média Diária(m³/s)"] * 1000
            df.insert(3, "Vazão Média Diária(l/s)", horario)

# Formatacao da data
            df["Data"] = pandas.to_datetime(df.Data)
            # Ano com Y maiúsculo, saída com 4 dígitos / Ano com y minúsculo, saída com 2 dígitos
            df["Data"] = df["Data"].dt.strftime("%d/%m/%Y")

# Carregar arquivo excel existente
            wb = load_workbook("C:/Users/tired/Desktop/" + Nome_Arquivo + ".xlsx")
            ws = wb["PN1-01"]

# Transformar dataframe em datarows (linhas de dados)
            dr = dataframe_to_rows(df, index=False, header=False,)

# Leitura de parâmetros do arquivo
            row_before = ws.max_row + 1
            column = ws.max_column

# Inserir dados na planilha
            for r in dr:
                ws.append(r)

# Inserir imagens na planilha
#            img1 = Image("C:/Users/Jair/Pictures/Acqua.png")
#            img2 = Image("C:/Users/Jair/Pictures/Lundin.png")

#            ws.add_image(img1, "A1")
#            ws.add_image(img2, "D1")

# Leitura de parâmetros do arquivo
            row_after = ws.max_row + 1

# Formatar dados da planilha
            for i in range(row_before, row_after):
                for j in range(1, column):
                    ws.cell(i, j).font = Font(name="Calibri",
                                              size=12)
#                                             bold = False,
#                                             italic = False,
                    ws.cell(i, j).border = Border(left=Side(border_style="thin",
                                                            color='FF000000'),
                                                  right=Side(border_style="thin",
                                                             color='FF000000'),
                                                  top=Side(border_style="thin",
                                                           color='FF000000'),
                                                  bottom=Side(border_style="thin",
                                                              color='FF000000'))
#                                                 diagonal=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 diagonal_direction=0,
#                                                 outline=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 vertical=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 horizontal=Side(border_style=None,
#                                                 color='FF000000'))

                    ws.cell(i, j).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(i, j).number_format = '0.00'

# Apresentacao dos dataframes no terminal
#            displayhook(df)

# Exportar dataframes como arquivo xlsx
#               df.to_excel("Teste Salvamento.xlsx", index= False) # Gerar arquivo pelo pandas
            wb.save("C:/Users/tired/Desktop/" + Nome_Salvar + ".xlsx")  # Gerar arquivo pelo openpyxl
            print("\nArquivo excel criado com sucesso!!!\n")

except OSError as e:
    print("Erro: ", e)
